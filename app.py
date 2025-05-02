from flask import Flask, request, jsonify, render_template, redirect, url_for, session
import joblib
import pandas as pd
import re
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from nltk.tokenize import word_tokenize
import os
import numpy as np
import sqlite3
from flask_mail import Mail, Message
from twilio.rest import Client
import random
import json
import requests
from openpyxl import Workbook, load_workbook
import re

# Get absolute path to this file's directory
base_dir = os.path.dirname(os.path.abspath(__file__))
template_dir = os.path.join(base_dir, 'templates')
model_dir = os.path.join(base_dir, 'models')

# Create Flask app with explicit template folder
app = Flask(__name__, template_folder=template_dir)

print(f"Current directory: {os.getcwd()}")
print(f"Base directory: {base_dir}")
print(f"Template directory: {template_dir}")
print(f"Model directory: {model_dir}")

def generate_verification_code():
    return random.randint(100000, 999999)

# Connect to the database
conn = sqlite3.connect('users.db')
cursor = conn.cursor()

# Create users table
# cursor.execute('''
# CREATE TABLE IF NOT EXISTS users (
#     id INTEGER PRIMARY KEY AUTOINCREMENT,
#     username TEXT NOT NULL,
#     password TEXT NOT NULL,
#     email TEXT NOT NULL,
#     mobile TEXT NOT NULL,
#     aadhaar TEXT NOT NULL
# )
# ''')

# Add a sample user (for testing purposes)
# cursor.execute('''
# INSERT INTO users (username, password, email, mobile, aadhaar)
# VALUES ('testuser', 'testpassword', 'testuser@example.com', '1234567890', '123456789012')
# ''')

cursor.execute('''
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL,
    password TEXT NOT NULL,
    aadhaar TEXT NOT NULL
)
''')
cursor.execute('''
INSERT INTO users (username, password, aadhaar)
VALUES ('testuser', 'testpassword', '123456789012')
''')
conn.commit()
conn.close()

# Method for generation of session key
# import secrets
# app.secret_key = secrets.token_hex(16)


# Secret key for session management
app.secret_key = 'e5b8c3a7d9f4e2a1b6c8d7e9f0a2b3c4'  # Replace with a secure random key

# Load the balanced model
try:
    model_path = os.path.join(model_dir, 'balanced_ipc_model.pkl')
    print(f"Loading model from: {model_path}")
    print(f"File exists: {os.path.exists(model_path)}")
    
    model_package = joblib.load(model_path)
    single_model = model_package['single_model']
    multi_model = model_package['multi_model']
    vectorizer = model_package['vectorizer']
    mlb = model_package['mlb']
    ipc_reference = model_package['ipc_reference']
    common_crime_sections = model_package.get('common_crime_sections', {})
    crime_related_sections = model_package.get('crime_related_sections', {})
    print("Balanced model loaded successfully!")
    
    # Define comprehensive crime keywords for direct detection
    crime_keywords = {
        "rape": ["rape", "raped", "sexual assault", "sexually assaulted", "forced himself", "against my will", "sexually abused"],
        "sexual_harassment": ["molest", "molested", "touched inappropriately", "sexual advances", "sexual harassment"],
        "robbery": ["robbed", "robbery", "snatched", "knife", "gun", "weapon", "threatened", "motorcycle", "bike"],
        "theft": ["stolen", "theft", "stole", "pickpocket", "took", "missing", "burglary", "broke into"],
        "assault": ["assault", "hit", "beat", "attacked", "injured", "hurt", "slapped", "punched"],
        "kidnapping": ["kidnap", "abduct", "ransom", "hostage", "forcibly took", "confined"],
        "murder": ["murder", "killed", "death", "deceased", "homicide", "died"],
        "fraud": ["fraud", "cheated", "deceived", "fake", "duped", "scam", "fraudulent"],
        "criminal_intimidation": ["threaten", "intimidate", "fear", "harm", "warning", "threatened"],
        "wrongful_confinement": ["locked", "confined", "detained", "held against will", "prevented from leaving"],
        "hurt": ["injury", "wound", "pain", "bleeding", "bruise", "physical harm"]
    }
    
except Exception as e:
    print(f"Error loading model: {str(e)}")
    # Provide placeholders for development
    single_model = None
    multi_model = None
    vectorizer = None
    mlb = None
    ipc_reference = {}
    common_crime_sections = {}
    crime_related_sections = {}
    crime_keywords = {}

# Ensure NLTK resources are downloaded
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt')

try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('stopwords')

try:
    nltk.data.find('corpora/wordnet')
except LookupError:
    nltk.download('wordnet')

# Enhanced text preprocessing function with comprehensive crime term preservation
def preprocess_text(text):
    """Preprocess text for machine learning with comprehensive crime term preservation"""
    if pd.isna(text):
        return ""
    
    # Convert to lowercase
    text = str(text).lower()
    
    # Preserve common legal phrases before removing punctuation
    text = text.replace("section ", "section_")
    
    # Replace specific crime terms with tokens
    crime_terms = {
        # Sexual offenses
        "rape": "rape_offense",
        "raped": "rape_offense",
        "sexual": "sexual_offense",
        "sexually": "sexual_offense",
        "molest": "sexual_offense",
        "molestation": "sexual_offense",
        "assault": "assault_offense",
        "assaulted": "assault_offense",
        "forced": "forced_act",
        "forcibly": "forced_act",
        "consent": "consent_issue",
        "against will": "consent_issue",
        "outrage": "outrage_modesty",
        "modesty": "outrage_modesty",
        
        # Robbery and theft related
        "robbed": "robbery_crime",
        "robbery": "robbery_crime",
        "snatched": "robbery_crime",
        "snatching": "robbery_crime",
        "knife": "weapon_knife",
        "knife-point": "weapon_knife",
        "gun": "weapon_gun",
        "pistol": "weapon_gun",
        "weapon": "weapon_used",
        "motorcycle": "vehicle_motorcycle",
        "bike": "vehicle_motorcycle",
        "money": "property_money",
        "cash": "property_money",
        "wallet": "property_wallet",
        "bag": "property_bag",
        "purse": "property_purse",
        "documents": "property_documents",
        "jewellery": "property_jewellery",
        "gold": "property_gold",
        "phone": "property_phone",
        "mobile": "property_phone",
        
        # Physical violence
        "hit": "physical_assault",
        "pushed": "physical_assault",
        "injured": "physical_injury",
        "injury": "physical_injury",
        "hurt": "physical_injury",
        "beaten": "physical_assault",
        "beat": "physical_assault",
        "attacked": "physical_assault",
        
        # Threats and intimidation
        "threatened": "criminal_intimidation",
        "threat": "criminal_intimidation",
        "intimidated": "criminal_intimidation",
        "fear": "criminal_intimidation",
        "warned": "criminal_intimidation",
        
        # Kidnapping
        "kidnapping": "kidnap_offense",
        "abduction": "kidnap_offense",
        "abducted": "kidnap_offense",
        "kidnapped": "kidnap_offense",
        "ransom": "kidnap_ransom",
        
        # Homicide
        "murder": "murder_offense",
        "killed": "murder_offense",
        "death": "murder_death",
        
        # Fraud and cheating
        "fraud": "fraud_offense",
        "cheated": "fraud_offense",
        "cheating": "fraud_offense",
        "forgery": "forgery_offense",
        "forged": "forgery_offense",
        
        # Other crimes
        "defamation": "defamation_offense",
        "trespass": "trespass_offense",
        "grievous": "grievous_hurt",
        "wrongful": "wrongful_act",
        "confinement": "confinement_offense",
        "restraint": "restraint_offense",
        "detained": "restraint_offense",
        "locked": "confinement_offense"
    }
    
    for term, replacement in crime_terms.items():
        text = re.sub(r'\b' + term + r'\b', replacement, text)
    
    # Remove special characters but keep spaces
    text = re.sub(r'[^\w\s]', ' ', text)
    
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Tokenize
    tokens = word_tokenize(text)
    
    # Remove stopwords (keeping legal terms)
    legal_stopwords = {'not', 'no', 'against', 'under', 'with', 'without', 'by', 'to', 'from', 
                      'between', 'within', 'upon', 'near', 'at', 'before', 'after'}
    stop_words = set(stopwords.words('english')) - legal_stopwords
    tokens = [word for word in tokens if word not in stop_words]
    
    # Lemmatization
    lemmatizer = WordNetLemmatizer()
    tokens = [lemmatizer.lemmatize(word) for word in tokens]
    
    return ' '.join(tokens)

# Function to detect crime types directly from text
def detect_crime_types(text):
    """Detect crime types directly from text using keywords"""
    text_lower = text.lower()
    detected_crimes = []
    
    for crime, keywords in crime_keywords.items():
        for keyword in keywords:
            if keyword in text_lower:
                detected_crimes.append(crime)
                break
    
    return detected_crimes

# Function to get sections based on detected crime types
def get_sections_by_crime(crime_types):
    """Get relevant IPC sections based on detected crime types"""
    sections = []
    for crime in crime_types:
        if crime in common_crime_sections:
            sections.extend(common_crime_sections[crime])
    
    return list(set(sections))  # Remove duplicates

# Function to get top sections with probabilities
def get_top_sections(model, features, classes, top_n=5, threshold=0.05):
    """Get top N section predictions with probabilities above threshold"""
    if hasattr(model, 'predict_proba'):
        probabilities = model.predict_proba(features)[0]
        # Get indices of top N probabilities
        top_indices = probabilities.argsort()[-top_n:][::-1]
        # Get sections and probabilities
        top_sections = [(classes[i], probabilities[i]) for i in top_indices if probabilities[i] > threshold]
        return top_sections
    else:
        # Fallback if predict_proba is not available
        prediction = model.predict(features)[0]
        return [(prediction, 1.0)]

def check_for_rape(text):
    """Check specifically for rape/sexual assault indicators"""
    text_lower = text.lower()
    
    # Check for rape indicators
    has_rape_term = any(term in text_lower for term in ["rape", "raped", "sexual assault", "sexually assaulted"])
    has_force = any(term in text_lower for term in ["forced", "forcibly", "against my will", "without consent"])
    has_threat = any(term in text_lower for term in ["threatened", "threat", "fear", "warned"])
    has_confinement = any(term in text_lower for term in ["locked", "confined", "detained", "couldn't leave"])
    has_physical = any(term in text_lower for term in ["hit", "hurt", "injury", "injured", "pushed", "assault"])
    
    # Score the rape likelihood
    score = 0
    if has_rape_term: score += 5
    if has_force: score += 3
    if has_threat: score += 2
    if has_confinement: score += 2
    if has_physical: score += 1
    
    # Determine sections based on score
    rape_sections = []
    if score >= 5:
        rape_sections.append(("Section 376", 0.95))  # Rape
        if has_confinement:
            rape_sections.append(("Section 342", 0.9))  # Wrongful confinement
        if has_threat:
            rape_sections.append(("Section 506", 0.9))  # Criminal intimidation
        if has_physical:
            rape_sections.append(("Section 323", 0.85))  # Voluntarily causing hurt
    
    return rape_sections, score >= 5

def check_for_robbery(text):
    """Check specifically for robbery indicators"""
    text_lower = text.lower()
    
    # Check for robbery indicators
    has_weapon = any(word in text_lower for word in ["knife", "gun", "pistol", "weapon"])
    has_taking = any(word in text_lower for word in ["snatched", "took", "robbed", "stole", "grabbed"])
    has_threat = any(word in text_lower for word in ["threatened", "threat", "fear", "forced"])
    has_vehicle = any(word in text_lower for word in ["motorcycle", "bike", "scooter", "vehicle"])
    has_money = any(word in text_lower for word in ["money", "cash", "wallet", "purse", "bag"])
    has_injury = any(word in text_lower for word in ["hit", "hurt", "injury", "injured", "push", "pushed", "assault"])
    
    # Score the robbery likelihood
    score = 0
    if has_weapon: score += 3
    if has_taking: score += 3
    if has_threat: score += 2
    if has_vehicle: score += 1
    if has_money: score += 1
    if has_injury: score += 2
    
    # Determine sections based on score
    robbery_sections = []
    if score >= 5:
        robbery_sections.append(("Section 392", 0.95))  # Robbery
        if has_injury:
            robbery_sections.append(("Section 394", 0.9))  # Voluntarily causing hurt in robbery
        if has_weapon:
            robbery_sections.append(("Section 397", 0.85))  # Robbery with deadly weapon
        if has_threat:
            robbery_sections.append(("Section 506", 0.8))  # Criminal intimidation
    
    return robbery_sections, score >= 5

@app.route('/')
def home():
    if 'logged_in' in session and session['logged_in']:
        return render_template('index.html')
    else:
        return redirect(url_for('login'))   

# @app.route('/login', methods=['GET', 'POST'])
# def login():
#     if request.method == 'POST':
#         # Get form data
#         username = request.form.get('username')
#         password = request.form.get('password')
#         aadhaar = request.form.get('aadhaar')

#         if not aadhaar.isdigit() or len(aadhaar) != 12:
#             return jsonify({'message': 'Invalid Aadhaar number. It must be 12 digits.'}), 400
        
#         # Connect to the database
#         conn = sqlite3.connect('users.db')
#         cursor = conn.cursor()

#         # Check if username and password match
#         cursor.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password))
#         user = cursor.fetchone()

#         # Check if Aadhaar matches
#         cursor.execute('SELECT * FROM users WHERE aadhaar = ?', (aadhaar,))
#         aadhaar_user = cursor.fetchone()

#         conn.close()

#         # Validate credentials
#         if user or aadhaar_user:
#             # Set session variable
#             session['logged_in'] = True
#             session['username'] = username
#             return jsonify({'message': 'Login successful! Redirecting...'}), 200
#         else:
#             return jsonify({'message': 'Invalid credentials, please try again.'}), 401
    
#     return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Get form data
        username = request.form.get('username')
        password = request.form.get('password')
        aadhaar = request.form.get('aadhaar')

        # Connect to the database
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()

        if password:
            password_regex = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$'
            if not re.match(password_regex, password):
                return jsonify({'message': 'Invalid password format. Please use a strong password.'}), 400


        # Case 1: Login with Aadhaar number only
        if aadhaar and not username and not password:
            if not aadhaar.isdigit() or len(aadhaar) != 12:
                return jsonify({'message': 'Invalid Aadhaar number. It must be 12 digits.'}), 400

            # Fetch the username associated with the Aadhaar number
            cursor.execute('SELECT username FROM users WHERE aadhaar = ?', (aadhaar,))
            user = cursor.fetchone()

            if user:
                username = user[0]  # Extract the username
                session['logged_in'] = True
                session['username'] = username
                conn.close()
                return jsonify({'message': f'Login successful! Welcome, {username}.'}), 200
            else:
                conn.close()
                return jsonify({'message': 'Invalid Aadhaar number. No user found.'}), 401

        # Case 2: Login with username and password
        elif username and password:
            cursor.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password))
            user = cursor.fetchone()

            if user:
                session['logged_in'] = True
                session['username'] = username
                conn.close()
                return jsonify({'message': f'Login successful! Welcome, {username}.'}), 200
            else:
                conn.close()
                return jsonify({'message': 'Invalid username or password.'}), 401

        # If neither case is satisfied
        else:
            conn.close()
            return jsonify({'message': 'Please provide either Aadhaar number or username and password.'}), 400

    return render_template('login.html')

@app.route('/signup', methods=['POST'])
def signup():
    try:
        # Get form data
        username = request.form.get('username')
        password = request.form.get('password')
        # email = request.form.get('email')
        # mobile = request.form.get('mobile')
        aadhaar = request.form.get('aadhaar')

        if not aadhaar.isdigit() or len(aadhaar) != 12:
            return jsonify({'message': 'Invalid Aadhaar number. It must be 12 digits.'}), 400

         # Validate password strength
        password_regex = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$'
        if not re.match(password_regex, password):
            return jsonify({'message': 'Password must contain at least one uppercase letter, one lowercase letter, one digit, one special character, and be at least 6 characters long.'}), 400
        
        # Connect to the database
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()

        # Check if the username or Aadhaar already exists
        cursor.execute('SELECT * FROM users WHERE username = ? OR aadhaar = ?', (username, aadhaar))
        existing_user = cursor.fetchone()

        if existing_user:
            return jsonify({'message': 'Username or Aadhaar already exists. Please try again.'}), 400

        # Insert the new user into the database
        # cursor.execute('''
        # INSERT INTO users (username, password, email, mobile, aadhaar)
        # VALUES (?, ?, ?, ?, ?)
        # ''', (username, password, email, mobile, aadhaar))
        cursor.execute('''
        INSERT INTO users (username, password, aadhaar)
        VALUES (?, ?, ?)
        ''', (username, password, aadhaar))
        conn.commit()
        conn.close()

        return jsonify({'message': 'Signup successful! You can now log in.'}), 200

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'message': 'An error occurred during signup. Please try again.'}), 500

@app.route('/logout')
def logout():
    session.clear()  # Clear all session data
    return redirect(url_for('login'))

@app.route('/predict', methods=['POST'])
def predict():
    try:
        # Get FIR text from request
        data = request.get_json()
        fir_text = data.get('fir_text', '')
        
        if not fir_text:
            return jsonify({'error': 'No FIR text provided'}), 400
        
        # Check if models are loaded
        if single_model is None or multi_model is None or vectorizer is None:
            return jsonify({'error': 'Models not loaded. Please check server logs.'}), 500
        
        # Detect crime types directly from text
        detected_crimes = detect_crime_types(fir_text)
        print(f"Detected crime types: {detected_crimes}")
        
        # Get sections based on detected crimes
        crime_based_sections = get_sections_by_crime(detected_crimes)
        print(f"Crime-based sections: {crime_based_sections}")
        
        # Special check for rape cases
        rape_sections, is_rape = check_for_rape(fir_text)
        print(f"Rape check: {is_rape}, Sections: {rape_sections}")
        
        # Special check for robbery cases
        robbery_sections, is_robbery = check_for_robbery(fir_text)
        print(f"Robbery check: {is_robbery}, Sections: {robbery_sections}")
        
        # Preprocess the text
        processed_text = preprocess_text(fir_text)
        
        # Vectorize the text
        text_features = vectorizer.transform([processed_text])
        
        # Get single-label predictions with probabilities
        single_sections = get_top_sections(
            single_model, 
            text_features, 
            single_model.classes_, 
            top_n=5, 
            threshold=0.05
        )

        # Get multi-label predictions
        multi_pred_binary = multi_model.predict(text_features)
        multi_pred_sections = mlb.inverse_transform(multi_pred_binary)[0]
        
        # Combine all section sources with priority
        all_sections_dict = {}
        
        # 1. Add rape-specific sections with highest priority if detected
        if is_rape:
            for section, prob in rape_sections:
                all_sections_dict[section] = max(prob, all_sections_dict.get(section, 0))
        
        # 2. Add robbery-specific sections with high priority if detected
        if is_robbery:
            for section, prob in robbery_sections:
                all_sections_dict[section] = max(prob, all_sections_dict.get(section, 0))
        
        # 3. Add crime-based sections with high priority
        for section in crime_based_sections:
            all_sections_dict[section] = max(0.85, all_sections_dict.get(section, 0))
        
        # 4. Add multi-label sections
        for section in multi_pred_sections:
            all_sections_dict[section] = max(0.75, all_sections_dict.get(section, 0))
        
        # 5. Add single-label sections
        for section, prob in single_sections:
            all_sections_dict[section] = max(prob, all_sections_dict.get(section, 0))
        
        # Convert to list and sort by confidence
        all_sections = [(section, prob) for section, prob in all_sections_dict.items()]
        all_sections.sort(key=lambda x: x[1], reverse=True)
        
        # Limit to top 5 sections
        top_sections = all_sections[:5]
        
        # Get section details for all predicted sections
        results = []
        for section, probability in top_sections:
            section_info = ipc_reference.get(section, {})
            if section_info:
                # Determine if section matches detected crimes
                section_category = section_info.get('category', '')
                
                results.append({
                    'section': section,
                    'description': section_info.get('description', 'Description not available'),
                    'offense': section_info.get('offense', 'Offense not available'),
                    'punishment': section_info.get('punishment', 'Punishment not available'),
                    'cognizable': section_info.get('cognizable', 'Not specified'),
                    'bailable': section_info.get('bailable', 'Not specified'),
                    'court': section_info.get('court', 'Not specified'),
                    'category': section_category,
                    'probability': f"{probability:.2%}"
                })
        
        # If no sections found, provide a fallback based on detected crimes
        if not results:
            if "rape" in detected_crimes or "sexual_harassment" in detected_crimes:
                fallback_sections = ["Section 376", "Section 354", "Section 506"]  # Default rape sections
            elif "robbery" in detected_crimes or "theft" in detected_crimes:
                fallback_sections = ["Section 392", "Section 379", "Section 506"]  # Default robbery/theft sections
            elif "assault" in detected_crimes or "hurt" in detected_crimes:
                fallback_sections = ["Section 323", "Section 324", "Section 506"]  # Default assault sections
            else:
                fallback_sections = []  # Generic fallback
                
            for section in fallback_sections:
                section_info = ipc_reference.get(section, {})
                if section_info:
                    results.append({
                        'section': section,
                        'description': section_info.get('description', 'Description not available'),
                        'offense': section_info.get('offense', 'Offense not available'),
                        'punishment': section_info.get('punishment', 'Punishment not available'),
                        'cognizable': section_info.get('cognizable', 'Not specified'),
                        'bailable': section_info.get('bailable', 'Not specified'),
                        'court': section_info.get('court', 'Not specified'),
                        'category': section_info.get('category', ''),
                        'probability': "70.00%"  # Default probability
                    })
        username = session.get('username')  # Get the logged-in user's username from the session
        if username:
            print(f"Saving to Excel for user: {username}")
            print(f"FIR Text: {fir_text}")
            print(f"IPC Sections: {all_sections}")
            save_to_excel(username, fir_text, list(all_sections))

        return jsonify({
            'success': True,
            'sections': results,
            'fir_text': fir_text,
            'detected_crimes': detected_crimes
        })
    
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/gemini', methods=['POST'])
def gemini_predict():
    try:
        # Get FIR text from request
        data = request.get_json()
        fir_text = data.get('fir_text', '')

        if not fir_text:
            return jsonify({'success': False, 'error': 'No FIR text provided'}), 400

        # Call Gemini API
        gemini_result = process_with_gemini(fir_text)

        if gemini_result.get('success'):
            ipc_sections = gemini_result.get('sections', [])
            username = session.get('username')  # Get the logged-in user's username from the session
            if username:
                print(f"Saving to Excel for user: {username}")
                print(f"FIR Text: {fir_text}")
                print(f"IPC Sections: {ipc_sections}")
                save_to_excel(username, fir_text, ipc_sections)

        # Directly return the result as JSON
        return jsonify(gemini_result)

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


def process_with_gemini(fir_text):
    try:
        gemini_api_url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"
        gemini_api_key = "AIzaSyA1OhwStr1ZIwsQc20asmfRRX38YAecGS0"  # Replace with your real API key

        # The prompt we send to Gemini API
        prompt = f"""
        You are an expert legal assistant. Given the following FIR complaint, predict which Indian Penal Code (IPC) sections apply.

        For each section, return:
        {{
            "section": "IPC 420",
            "category": "Cheating",
            "offense": "Cheating and dishonestly inducing delivery of property",
            "description": "Obtaining property by deception",
            "punishment": "Up to 7 years imprisonment and fine",
            "court": "Magistrate Court",
            "confidence": 85
        }}

        Also, give a list of detected crime categories (like ["Theft", "Assault"]).

        FIR: {fir_text}

        Format the full response as:
        {{
            "sections": [ ... ],
            "detected_crimes": [ ... ]
        }}
        """

        # Define the JSON payload to match the curl format
        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": prompt
                        }
                    ]
                }
            ]
        }

        # Define the headers to match the curl format
        headers = {
            "Content-Type": "application/json"
        }

        # Print payload and headers for debugging
        print("Sending request to Gemini API with payload:", json.dumps(payload, indent=2))
        print("Using headers:", headers)

        # Make the API request
        response = requests.post(
            f"{gemini_api_url}?key={gemini_api_key}",
            headers=headers,
            json=payload,
            timeout=30  # 30 seconds timeout
        )

        # Print response status code and content for debugging
        print("Response Status Code:", response.status_code)
        print("Response Text:", response.text)

        # Check if the response is successful (HTTP Status 200)
        if response.status_code == 200:
            data = response.json()

            # Extract the response text
            candidates = data.get('candidates', [])
            if not candidates:
                return {'success': False, 'error': 'No candidates returned by Gemini API'}

            result_text = candidates[0].get('content', {}).get('parts', [{}])[0].get('text', '')
            if not result_text:
                return {'success': False, 'error': 'No content in Gemini API response'}

            # Format the result for the frontend
            result = {
                'success': True,
                'fir_text': fir_text,
                'predictions': result_text
            }

            return result
        else:
            # Handle response failure
            return {
                'success': False,
                'error': 'Failed to process with Gemini API',
                'details': response.text
            }

    except Exception as e:
        return {
            'success': False,
            'error': 'An error occurred while using Gemini API',
            'details': str(e)
        }


def save_to_excel(username, fir_text, ipc_sections):
    """Save FIR complaint and IPC sections to an Excel file under the logged-in user."""
    try:
        # Define the file path for the user's Excel file
        file_path = f"{username}_fir_data.xlsx"

        # Extract only the section names if ipc_sections contains tuples
        ipc_sections = [section if isinstance(section, str) else section[0] for section in ipc_sections]

        # Check if the file already exists
        if os.path.exists(file_path):
            # Load the existing workbook
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            # Create a new workbook and add headers
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["FIR Complaint", "Predicted IPC Sections"])  # Add headers

        # Add the FIR complaint and IPC sections to the sheet
        sheet.append([fir_text, ", ".join(ipc_sections)])

        # Save the workbook
        workbook.save(file_path)
        print(f"Excel file saved successfully at {file_path}")

    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
# Test endpoint for specific crime types
@app.route('/test-case', methods=['POST'])
def test_case():
    """Test endpoint for crime detection"""
    data = request.get_json()
    test_case = data.get('fir_text', '')
    
    if not test_case:
        return jsonify({'error': 'No test case provided'}), 400
    
    # Process with our detection logic
    detected_crimes = detect_crime_types(test_case)
    rape_sections, is_rape = check_for_rape(test_case)
    robbery_sections, is_robbery = check_for_robbery(test_case)
    
    # Preprocess and get model predictions
    processed_text = preprocess_text(test_case)
    text_features = vectorizer.transform([processed_text])
    
    # Get single-label prediction
    if single_model:
        single_pred = single_model.predict(text_features)[0]
        
        # Get top 5 predictions if possible
        top_predictions = []
        if hasattr(single_model, 'predict_proba'):
            proba = single_model.predict_proba(text_features)[0]
            top_indices = proba.argsort()[-5:][::-1]
            top_classes = single_model.classes_[top_indices]
            top_probas = proba[top_indices]
            top_predictions = list(zip(top_classes, top_probas))
    else:
        single_pred = "Model not loaded"
        top_predictions = []
    
    # Get multi-label predictions if model is loaded
    if multi_model and mlb:
        multi_pred_binary = multi_model.predict(text_features)
        multi_pred_sections = mlb.inverse_transform(multi_pred_binary)[0]
    else:
        multi_pred_sections = []
    
    return jsonify({
        'test_case': test_case[:100] + "...",
        'detected_crimes': detected_crimes,
        'is_rape': is_rape,
        'rape_sections': rape_sections,
        'is_robbery': is_robbery,
        'robbery_sections': robbery_sections,
        'single_prediction': single_pred,
        'top_predictions': [{"section": s, "probability": float(p)} for s, p in top_predictions],
        'multi_label_predictions': list(multi_pred_sections)
    })

if __name__ == '__main__':
    app.run(debug=True, port=5050)